"""
Export a normalized rates DataFrame to a matrix-layout Excel workbook.

Shipment detail columns appear first; each rate_column becomes a two-column
cost block (Currency + Flat or p/unit) with a multi-row header.
"""

from __future__ import annotations

import re
import sys
from collections import OrderedDict
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parent
PROCESSING_DIR = PROJECT_ROOT / "processing"
OUTPUT_DIR = PROJECT_ROOT / "output"

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from common import (
    format_p_unit_label,
    is_blank,
    is_flat_charge_label,
    is_flat_price_per,
    is_roundtrip_rate_group,
    is_truthy,
    ltl_transport_block_title,
    normalize_header,
)

# Shipment columns: internal name -> Excel header (only if column exists in df)
LANE_NUM_HEADER = "Lane #"
TAB_HEADER = "Tab"

SHIPMENT_COLUMN_MAP: OrderedDict[str, str] = OrderedDict(
    [
        ("lane_id", "Lane ID (Milkrun ID)"),
        ("lane_description", "Lane Description"),
        ("origin_country", "Origin Country"),
        ("origin_zip", "Origin Postal Code"),
        ("origin_city", "Origin City"),
        ("dest_country", "Destination Country"),
        ("dest_zip", "Destination Postal Code"),
        ("dest_city", "Destination City"),
        ("rate_group", "Service"),
        ("equipment_type", "Equipment Type"),
        ("price_per", "Price Per"),
    ]
)

# Included only when the column exists and has at least one non-blank value
CONDITIONAL_SHIPMENT_COLUMNS = frozenset(
    {"equipment_type", "price_per", "lane_description"}
)

# Optional extra mappings — extend SHIPMENT_COLUMN_MAP or pass shipment_columns=
OPTIONAL_SHIPMENT_COLUMN_MAP: OrderedDict[str, str] = OrderedDict(
    [
        ("origin_name", "Carrier"),
        ("paid_by", "Paid by"),
        ("dest_name", "Destination name"),
        ("cost_component", "Cost component"),
        ("description", "Description"),
    ]
)

ROUNDTRIP_RATE_COLUMN = "Roundtrip"

LTL_MIN_RATE = "Minimum Rate"
LTL_MAX_RATE = "Maximum Rate"
LTL_ROUND_TRIP = "Round Trip?"

WEIGHT_RANGE_RE = re.compile(r"^(\d+)\s*-\s*(\d+)$")
WEIGHT_KG_COLUMN_RE = re.compile(r"^(\d+(?:[.,]\d+)?)\s*kg\s*$", re.I)
TILL_KG_RE = re.compile(r"^till\s*([\d.,]+)\s*kg\s*$", re.I)
KG_RANGE_HEADER_RE = re.compile(
    r"^([\d.,]+)\s*kg\s*-\s*([\d.,]+)\s*kg\s*$", re.I
)
SPURIOUS_KG_RANGE_RE = re.compile(r"^(\d+)\s*kg\s*-\s*(\d+)\s*kg\s*$", re.I)
PALLET_RANGE_RE = re.compile(r"^(\d+)\s*-\s*(\d+)\s+pallet", re.I)
PALLET_COUNT_RE = re.compile(r"^\d{1,3}$")
PALLET_PLT_RE = re.compile(r"^(\d+)\s*plt\s*$", re.I)
MAUT_RATE_COLUMN_RE = re.compile(r"\bmaut\b", re.I)

EMONS_FLAT_MIN_KG = "100 kg"
EMONS_FLAT_MAX_RC = "FTL"

META_COLUMNS = frozenset(
    {
        "layout",
        "source_file",
        "sheet_name",
        "row_number",
        "rate_column",
        "price",
        "currency",
        "cost_component",
        "description",
        "lane_number",
        "roundtrip",
        "equipment_type",
        "price_per",
    }
)

WEIGHT_BAND_RE = re.compile(
    r"\bkg\b|till\s*[\d.,]+|^\d+[\d.,]*\s*-\s*\d|^[\d,.\s]+$",
    re.I,
)

HEADER_ROW_COUNT = 4
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
BOLD = Font(bold=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _is_weight_band(rate_column: str) -> bool:
    t = normalize_header(rate_column)
    if not t:
        return False
    if WEIGHT_BAND_RE.search(str(rate_column).strip()):
        return True
    if re.fullmatch(r"[\d,.\s]+", str(rate_column).replace(",", "").strip()):
        return True
    return False


def _cost_block_title(rate_column: str) -> str:
    return f"Transport cost ({rate_column})"


def _ltl_band_upper_bound(rate_column: str) -> int | None:
    m = WEIGHT_RANGE_RE.match(str(rate_column).strip())
    if not m:
        return None
    return int(m.group(2))


def _ltl_band_header(rate_column: str) -> str | None:
    upper = _ltl_band_upper_bound(rate_column)
    if upper is None:
        return None
    return f"<={upper}"


def _parse_kg_number(text: str) -> int:
    s = str(text).strip().replace(" ", "")
    if re.fullmatch(r"\d{1,3}\.\d{3}", s):
        s = s.replace(".", "")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")
    return int(float(s))


def _is_groupage_kg_rate_column(rate_column: str) -> bool:
    text = str(rate_column).strip()
    return bool(TILL_KG_RE.match(text) or KG_RANGE_HEADER_RE.match(text))


def _groupage_kg_upper_bound(rate_column: str) -> int | None:
    text = str(rate_column).strip()
    m = TILL_KG_RE.match(text)
    if m:
        return _parse_kg_number(m.group(1))
    m = KG_RANGE_HEADER_RE.match(text)
    if m:
        return _parse_kg_number(m.group(2))
    return None


def _is_spurious_kg_range(rate_column: str) -> bool:
    m = SPURIOUS_KG_RANGE_RE.match(str(rate_column).strip())
    if not m:
        return False
    return m.group(1) == m.group(2) and int(m.group(1)) < 200


def _is_simple_kg_rate_column(rate_column: str) -> bool:
    text = str(rate_column).strip()
    if _is_groupage_kg_rate_column(text) or _is_spurious_kg_range(text):
        return False
    if WEIGHT_KG_COLUMN_RE.match(text):
        return True
    if WEIGHT_RANGE_RE.match(text):
        return True
    if _ltl_band_upper_bound(text) is not None:
        return True
    return False


def _is_kg_bracket_rate_column(rate_column: str) -> bool:
    text = str(rate_column).strip()
    if _is_simple_kg_rate_column(text):
        return True
    if _is_groupage_kg_rate_column(text):
        return True
    return False


def _is_maut_rate_column(rate_column: str) -> bool:
    return bool(MAUT_RATE_COLUMN_RE.search(normalize_header(str(rate_column).strip())))


def _df_without_maut(df: pd.DataFrame) -> pd.DataFrame:
    if "rate_column" not in df.columns:
        return df
    mask = ~df["rate_column"].astype(str).apply(_is_maut_rate_column)
    return df.loc[mask].copy()


def _is_pallet_bracket_rate_column(rate_column: str) -> bool:
    text = str(rate_column).strip()
    if normalize_header(text) == "ftl":
        return False
    if _is_maut_rate_column(text):
        return False
    if PALLET_PLT_RE.match(text):
        return True
    if PALLET_RANGE_RE.match(text):
        return True
    if PALLET_COUNT_RE.match(text):
        return int(text) <= 100
    return False


def _pallet_band_display_header(rate_column: str) -> str | None:
    text = str(rate_column).strip()
    m = PALLET_PLT_RE.match(text)
    if m:
        return f"<={m.group(1)}"
    m = PALLET_RANGE_RE.match(text)
    if m:
        return f"<={m.group(2)}"
    if PALLET_COUNT_RE.match(text):
        return f"<={text}"
    return None


def _parse_kg_band_value(rate_column: str) -> int | None:
    text = str(rate_column).strip()
    m = WEIGHT_KG_COLUMN_RE.match(text)
    if m:
        return int(float(m.group(1).replace(",", "")))
    return None


def _kg_bracket_columns_ordered(df: pd.DataFrame) -> list[str]:
    seen: list[str] = []
    for val in df["rate_column"]:
        if is_blank(val):
            continue
        text = str(val).strip()
        if _is_simple_kg_rate_column(text) and text not in seen:
            seen.append(text)
    seen.sort(key=lambda c: _parse_kg_band_value(c) or 0)
    return seen


def _groupage_kg_columns_ordered(df: pd.DataFrame) -> list[str]:
    seen: list[str] = []
    for val in df["rate_column"]:
        if is_blank(val):
            continue
        text = str(val).strip()
        if _is_groupage_kg_rate_column(text) and text not in seen:
            seen.append(text)
    seen.sort(key=lambda c: _groupage_kg_upper_bound(c) or 0)
    return seen


def _groupage_band_display_header(rate_column: str) -> str | None:
    upper = _groupage_kg_upper_bound(rate_column)
    if upper is None:
        return None
    return f"<={upper}"


def _pallet_bracket_columns_ordered(df: pd.DataFrame) -> list[str]:
    seen: list[str] = []
    for val in df["rate_column"]:
        if is_blank(val):
            continue
        text = str(val).strip()
        if _is_pallet_bracket_rate_column(text) and text not in seen:
            seen.append(text)
    seen.sort(key=lambda c: _pallet_band_sort_value(c))
    return seen


def _pallet_band_sort_value(rate_column: str) -> int:
    hdr = _pallet_band_display_header(rate_column)
    if hdr:
        m = re.search(r"(\d+)", hdr)
        if m:
            return int(m.group(1))
    return 0


def _kg_bands_need_ccc_ceilings(kg_bands: list[str]) -> bool:
    """True when band labels use CCC-style sub-header ceilings (50 kg -> <=50, 100 kg -> <=101)."""
    values = [_parse_kg_band_value(b) for b in kg_bands]
    values = [v for v in values if v is not None]
    if len(values) < 2:
        return False
    for i in range(len(values) - 1):
        if values[i + 1] >= 4 * values[i]:
            return True
    return False


def _kg_band_display_header(rate_column: str, kg_bands: list[str]) -> str | None:
    """Kg ceiling label: groupage till/range, simple <=N, or CCC-style when detected."""
    groupage = _groupage_band_display_header(rate_column)
    if groupage:
        return groupage
    n = _parse_kg_band_value(rate_column)
    if n is None:
        return None
    if not _kg_bands_need_ccc_ceilings(kg_bands):
        return f"<={n}"
    text = str(rate_column).strip()
    try:
        i = kg_bands.index(text)
    except ValueError:
        return f"<={n}"
    if i == 0 or i == len(kg_bands) - 1:
        return f"<={n}"
    next_n = _parse_kg_band_value(kg_bands[i + 1])
    if next_n is None:
        return f"<={n}"
    if next_n < 10 * n:
        return f"<={n + 1}"
    return f"<={next_n + 1}"


def _bracket_families(df: pd.DataFrame) -> list[tuple[str, list[str], str]]:
    """(family_key, ordered rate_column bands, rate-by category for row 3)."""
    families: list[tuple[str, list[str], str]] = []
    kg = _kg_bracket_columns_ordered(df)
    if kg:
        families.append(("kg", kg, "Weight/kg"))
    groupage = _groupage_kg_columns_ordered(df)
    if groupage:
        families.append(("kg_groupage", groupage, "Weight/kg"))
    pal = _pallet_bracket_columns_ordered(df)
    if pal:
        families.append(("pallet", pal, "Pallet"))
    return families


def _is_transport_price_per(value: Any) -> bool:
    if is_blank(value):
        return False
    t = normalize_header(str(value).strip())
    if t in (
        "flat rate",
        "minimum (flat rate)",
        "maximum (flat rate)",
        "ltl standard",
    ):
        return True
    if "per 100" in t and "kg" in t:
        return True
    if re.search(r"per\s+\d+", t) and "kg" in t:
        return True
    return False


def _is_dual_charge_minmax_ltl(df: pd.DataFrame) -> bool:
    if not _is_ltl_transport_grid(df):
        return False
    return bool(_kg_bands_for_price_per(df, "flat rate")) and bool(
        _kg_bands_for_price_per(df, "per 100 kg")
    )


def _is_min_rate_column(rate_column: str) -> bool:
    t = normalize_header(str(rate_column).strip())
    return t in (
        "minimum rate",
        "minimum (flat rate)",
        "min price",
        "min. charge",
        "min price",
    )


def _is_max_rate_column(rate_column: str) -> bool:
    t = normalize_header(str(rate_column).strip())
    return t in ("maximum rate", "maximum (flat rate)", "max price")


def _is_weight_bracket_rate_column(rate_column: str) -> bool:
    if is_blank(rate_column):
        return False
    text = str(rate_column).strip()
    if _is_min_rate_column(text) or _is_max_rate_column(text):
        return True
    if _is_kg_bracket_rate_column(text) or _is_pallet_bracket_rate_column(text):
        return True
    if _ltl_band_header(text):
        return True
    return False


def _weight_rate_column_key(rate_column: str) -> str:
    text = str(rate_column).strip()
    if _is_min_rate_column(text):
        return LTL_MIN_RATE
    if _is_max_rate_column(text):
        return LTL_MAX_RATE
    return text


def _weight_band_display_header(
    rate_column: str,
    *,
    kg_bands: list[str] | None = None,
) -> str | None:
    if _is_min_rate_column(rate_column) or _is_max_rate_column(rate_column):
        return None
    pallet = _pallet_band_display_header(rate_column)
    if pallet:
        return pallet
    text = str(rate_column).strip()
    if _is_kg_bracket_rate_column(text):
        bands = kg_bands if kg_bands is not None else _kg_bracket_columns_ordered_from_rate(text)
        return _kg_band_display_header(text, bands) if bands else f"<={_parse_kg_band_value(text)}"
    band = _ltl_band_header(rate_column)
    if band:
        return band
    return None


def _kg_bracket_columns_ordered_from_rate(rate_column: str) -> list[str]:
    """Single-column fallback when kg_bands list is not provided."""
    n = _parse_kg_band_value(rate_column)
    return [rate_column] if n is not None else []


def _weight_band_sort_value(rate_column: str) -> int:
    if _is_pallet_bracket_rate_column(rate_column):
        return _pallet_band_sort_value(rate_column)
    upper = _groupage_kg_upper_bound(rate_column)
    if upper is not None:
        return upper
    n = _parse_kg_band_value(rate_column)
    if n is not None:
        return n
    header = _ltl_band_header(rate_column)
    if header:
        m = re.search(r"(\d+)", header)
        return int(m.group(1)) if m else 0
    return 0


def _is_ltl_transport_grid(df: pd.DataFrame) -> bool:
    if "rate_column" not in df.columns:
        return False
    cols = [str(c).strip() for c in df["rate_column"].dropna().unique()]
    has_min = any(_is_min_rate_column(c) or c == LTL_MIN_RATE for c in cols)
    has_max = any(_is_max_rate_column(c) or c == LTL_MAX_RATE for c in cols)
    has_band = bool(_kg_bracket_columns_ordered(df)) or any(
        _ltl_band_header(c) for c in cols
    )
    return has_min and has_max and has_band


def _kg_bands_for_price_per(df: pd.DataFrame, price_per_norm: str) -> list[str]:
    seen: list[str] = []
    if "price_per" not in df.columns:
        return seen
    for _, row in df.iterrows():
        rc = str(row.get("rate_column", "")).strip()
        if not _is_simple_kg_rate_column(rc) or normalize_header(rc) == "ftl":
            continue
        pp = normalize_header(row.get("price_per"))
        if pp != price_per_norm:
            continue
        if rc not in seen:
            seen.append(rc)
    seen.sort(key=lambda c: _parse_kg_band_value(c) or 0)
    return seen


def _groupage_bands_for_price_per(df: pd.DataFrame, price_per_norm: str) -> list[str]:
    seen: list[str] = []
    if "price_per" not in df.columns:
        return seen
    for _, row in df.iterrows():
        rc = str(row.get("rate_column", "")).strip()
        if not _is_groupage_kg_rate_column(rc):
            continue
        pp = normalize_header(row.get("price_per"))
        if pp != price_per_norm:
            continue
        if rc not in seen:
            seen.append(rc)
    seen.sort(key=lambda c: _groupage_kg_upper_bound(c) or 0)
    return seen


def _price_per_label_for_norm(df: pd.DataFrame, price_per_norm: str) -> str:
    for val in _ltl_price_per_values(df):
        if normalize_header(val) == price_per_norm:
            return val
    return price_per_norm


def _is_emons_synthetic_min_max_grid(df: pd.DataFrame) -> bool:
    """FLAT RATE 100 kg = MIN, FLAT RATE FTL = MAX, PER 100 KG on other kg bands."""
    if "price_per" not in df.columns or not _kg_bracket_columns_ordered(df):
        return False
    if _is_ltl_transport_grid(df):
        return False
    pps = {normalize_header(str(p)) for p in df["price_per"].dropna() if not is_blank(p)}
    if "flat rate" not in pps:
        return False
    if not any("per 100" in p for p in pps):
        return False
    sub = df[df["rate_column"].astype(str).str.strip() == EMONS_FLAT_MIN_KG]
    has_100_flat = bool(
        sub["price_per"]
        .astype(str)
        .apply(lambda x: normalize_header(x) == "flat rate")
        .any()
    )
    sub_ftl = df[df["rate_column"].astype(str).str.strip().str.upper() == EMONS_FLAT_MAX_RC]
    has_ftl_flat = bool(
        sub_ftl["price_per"]
        .astype(str)
        .apply(lambda x: normalize_header(x) == "flat rate")
        .any()
    )
    return has_100_flat and has_ftl_flat


def _is_emons_ltl_grid(df: pd.DataFrame) -> bool:
    return _is_emons_synthetic_min_max_grid(df) and (
        len(_pallet_bracket_columns_ordered(df)) >= 2
        or bool(_kg_bands_for_price_per(df, "per 100 kg"))
    )


def _emons_flat_cost_specs(df: pd.DataFrame) -> list[tuple[str, str, str]]:
    bands = _kg_bands_for_price_per(df, "flat rate")
    kg_bands = bands
    charge = _price_per_label_for_norm(df, "flat rate")
    specs: list[tuple[str, str, str]] = [("__currency__", "Currency", "")]
    specs.append((LTL_MIN_RATE, "MIN", charge))
    for band in bands:
        header = _kg_band_display_header(band, kg_bands) or band
        specs.append((band, header, charge))
    specs.append((LTL_MAX_RATE, "MAX", charge))
    return specs


def _emons_per100_cost_specs(df: pd.DataFrame) -> list[tuple[str, str, str]]:
    bands = _kg_bands_for_price_per(df, "per 100 kg")
    kg_bands = bands
    charge = _price_per_label_for_norm(df, "per 100 kg")
    specs: list[tuple[str, str, str]] = [("__currency__", "Currency", "")]
    for band in bands:
        header = _kg_band_display_header(band, kg_bands) or band
        specs.append((band, header, charge))
    return specs


def _ltl_minmax_flat_cost_specs(df: pd.DataFrame) -> list[tuple[str, str, str]]:
    bands = _kg_bands_for_price_per(df, "flat rate")
    charge = _price_per_label_for_norm(df, "flat rate")
    specs: list[tuple[str, str, str]] = [("__currency__", "Currency", "")]
    specs.append(
        (
            LTL_MIN_RATE,
            "MIN",
            _charge_label_for_rate_key(df, LTL_MIN_RATE) or charge,
        )
    )
    for band in bands:
        header = _kg_band_display_header(band, bands) or band
        specs.append((band, header, charge))
    specs.append(
        (
            LTL_MAX_RATE,
            "MAX",
            _charge_label_for_rate_key(df, LTL_MAX_RATE) or charge,
        )
    )
    return specs


def _ltl_per100_only_cost_specs(df: pd.DataFrame) -> list[tuple[str, str, str]]:
    bands = _kg_bands_for_price_per(df, "per 100 kg")
    charge = _price_per_label_for_norm(df, "per 100 kg")
    specs: list[tuple[str, str, str]] = [("__currency__", "Currency", "")]
    for band in bands:
        header = _kg_band_display_header(band, bands) or band
        specs.append((band, header, charge))
    return specs


def _df_for_matrix_export(df: pd.DataFrame) -> pd.DataFrame:
    """Drop MAUT price columns and Total cost rows before matrix export."""
    df = _df_without_maut(df)
    if "cost_component" in df.columns:
        mask = ~df["cost_component"].astype(str).apply(
            lambda x: normalize_header(x) == "total cost"
        )
        df = df.loc[mask].copy()
    elif "description" in df.columns:
        mask = ~df["description"].astype(str).apply(
            lambda x: normalize_header(x) == "total cost"
        )
        df = df.loc[mask].copy()
    return df


def _cost_component_values(df: pd.DataFrame) -> list[str]:
    """Rate / Maut row types for cost-component-split export (excludes Total cost)."""
    if "cost_component" not in df.columns:
        return []
    by_norm: dict[str, str] = {}
    for val in df["cost_component"]:
        if is_blank(val):
            continue
        text = str(val).strip()
        norm = normalize_header(text)
        if norm in ("total cost", "total"):
            continue
        if norm in ("rate", "maut") and norm not in by_norm:
            by_norm[norm] = text
    order = {"rate": 0, "maut": 1}
    return sorted(by_norm.values(), key=lambda x: order.get(normalize_header(x), 9))


def _is_cost_component_ltl_grid(df: pd.DataFrame) -> bool:
    """Dual-charge MIN/MAX grid with separate Rate and Maut rows (e.g. Ceva HU)."""
    if not _is_dual_charge_minmax_ltl(df):
        return False
    comps = {normalize_header(c) for c in _cost_component_values(df)}
    return "rate" in comps and "maut" in comps


def _combined_minmax_cost_specs(df: pd.DataFrame) -> list[tuple[str, str, str]]:
    """One block: MIN, flat kg bands, PER 100 kg bands, MAX (mixed column suffixes)."""
    flat_bands = _kg_bands_for_price_per(df, "flat rate")
    per100_bands = _kg_bands_for_price_per(df, "per 100 kg")
    flat_charge = _price_per_label_for_norm(df, "flat rate")
    per100_charge = _price_per_label_for_norm(df, "per 100 kg")
    specs: list[tuple[str, str, str]] = [("__currency__", "Currency", "")]
    specs.append(
        (
            LTL_MIN_RATE,
            "MIN",
            _charge_label_for_rate_key(df, LTL_MIN_RATE) or flat_charge,
        )
    )
    for band in flat_bands:
        header = _kg_band_display_header(band, flat_bands) or band
        specs.append((band, header, flat_charge))
    for band in per100_bands:
        header = _kg_band_display_header(band, per100_bands) or band
        specs.append((band, header, per100_charge))
    specs.append(
        (
            LTL_MAX_RATE,
            "MAX",
            _charge_label_for_rate_key(df, LTL_MAX_RATE) or flat_charge,
        )
    )
    return specs


def _row_cost_component_norm(row: pd.Series) -> str:
    for col in ("cost_component", "description"):
        if col in row.index and not is_blank(row.get(col)):
            return normalize_header(str(row.get(col)).strip())
    return ""


def _extract_combined_minmax_prices(
    group: pd.DataFrame,
    cost_specs: list[tuple[str, str, str]],
    cost_component: str | None = None,
) -> dict[str, Any]:
    """Combined flat + per-100-kg block; optional filter by Rate/Maut cost_component."""
    cc_norm = normalize_header(cost_component) if cost_component else None
    flat_bands = [
        key
        for key, _, charge in cost_specs
        if key not in ("__currency__", LTL_MIN_RATE, LTL_MAX_RATE)
        and is_flat_charge_label(charge)
    ]
    per100_bands = [
        key
        for key, _, charge in cost_specs
        if key not in ("__currency__", LTL_MIN_RATE, LTL_MAX_RATE)
        and not is_flat_charge_label(charge)
    ]
    prices: dict[str, Any] = {}
    currency: Any = None
    for _, row in group.iterrows():
        if cc_norm is not None and _row_cost_component_norm(row) != cc_norm:
            continue
        rc = str(row["rate_column"]).strip()
        pp = normalize_header(row.get("price_per"))
        if currency is None and not is_blank(row.get("currency")):
            currency = row.get("currency")
        price = row.get("price")
        if _is_min_rate_column(rc) or pp == "minimum (flat rate)":
            prices[LTL_MIN_RATE] = price
        elif _is_max_rate_column(rc) or pp == "maximum (flat rate)":
            prices[LTL_MAX_RATE] = price
        elif rc in flat_bands and pp == "flat rate":
            prices[rc] = price
        elif rc in per100_bands and "per 100" in pp:
            prices[rc] = price
    prices["__currency__"] = currency
    return prices


def _groupage_cost_specs(
    df: pd.DataFrame,
    bands: list[str],
) -> list[tuple[str, str, str]]:
    charge = _price_per_label_for_norm(df, "per 100 kg")
    specs: list[tuple[str, str, str]] = [("__currency__", "Currency", "")]
    for band in bands:
        header = _groupage_band_display_header(band) or band
        specs.append((band, header, charge))
    return specs


def _bracket_transport_block_specs(
    df: pd.DataFrame,
) -> list[tuple[str, str, list[tuple[str, str, str]]]]:
    """(family_key, price_per block key, cost_specs) for all transport bracket blocks."""
    if _is_cost_component_ltl_grid(df):
        specs = _combined_minmax_cost_specs(df)
        return [
            ("kg_combined", comp, specs) for comp in _cost_component_values(df)
        ]

    if _is_emons_synthetic_min_max_grid(df):
        blocks: list[tuple[str, str, list[tuple[str, str, str]]]] = []
        pal = _pallet_bracket_columns_ordered(df)
        if pal:
            blocks.append(
                ("pallet", "", _bracket_cost_specs(df, pal, "pallet"))
            )
        flat_pp = _price_per_label_for_norm(df, "flat rate")
        blocks.append(("kg_flat", flat_pp, _emons_flat_cost_specs(df)))
        per100_pp = _price_per_label_for_norm(df, "per 100 kg")
        blocks.append(("kg_per100", per100_pp, _emons_per100_cost_specs(df)))
        return blocks

    blocks: list[tuple[str, str, list[tuple[str, str, str]]]] = []
    if _is_dual_charge_minmax_ltl(df):
        specs = _combined_minmax_cost_specs(df)
        blocks.append(("kg_combined", "", specs))
    elif _is_ltl_transport_grid(df):
        flat_pp = _price_per_label_for_norm(df, "flat rate") or ""
        blocks.append(("kg", flat_pp, _ltl_cost_specs(df)))

    groupage = _groupage_bands_for_price_per(df, "per 100 kg")
    if not groupage:
        groupage = _groupage_kg_columns_ordered(df)
    if groupage:
        per100_pp = _price_per_label_for_norm(df, "per 100 kg")
        blocks.append(
            ("kg_groupage", per100_pp, _groupage_cost_specs(df, groupage))
        )

    if blocks:
        return blocks

    price_per_values = _ltl_price_per_values_ordered(df)
    if not price_per_values:
        price_per_values = [""]
    for family_key, bands, _rate_by in _bracket_families(df):
        if family_key == "pallet":
            blocks.append(
                ("pallet", "", _bracket_cost_specs(df, bands, family_key))
            )
        elif family_key == "kg_groupage":
            per100_pp = _price_per_label_for_norm(df, "per 100 kg")
            blocks.append(
                (
                    "kg_groupage",
                    per100_pp,
                    _groupage_cost_specs(df, bands),
                )
            )
        else:
            for pp in price_per_values:
                blocks.append(
                    (family_key, pp, _bracket_cost_specs(df, bands, family_key))
                )
    return blocks


def _bracket_block_title(family_key: str, price_per: str) -> str:
    if family_key == "kg_combined":
        if price_per:
            return f"Transport cost ({price_per})"
        return "Transport cost"
    if family_key == "pallet":
        return "Transport cost (Flat)"
    if family_key in ("kg_flat", "kg"):
        return "Transport cost (Flat)"
    if family_key in ("kg_per100", "kg_groupage"):
        return f"Transport cost ({format_p_unit_label(price_per)})"
    return ltl_transport_block_title(price_per)


def _bracket_rate_by_category(family_key: str) -> str:
    if family_key == "pallet":
        return "Pallet"
    return "Weight/kg"


def _is_open_bracket_transport_grid(df: pd.DataFrame) -> bool:
    """Kg and/or pallet brackets without requiring MIN+MAX (e.g. CCC LTL pallets / LTL CC)."""
    if _is_emons_synthetic_min_max_grid(df):
        return False
    if _is_dual_charge_minmax_ltl(df):
        return True
    if _is_ltl_transport_grid(df):
        return True
    groupage = _groupage_kg_columns_ordered(df)
    if len(groupage) >= 2:
        return True
    kg = _kg_bracket_columns_ordered(df)
    pal = _pallet_bracket_columns_ordered(df)
    if len(kg) >= 2:
        return True
    if len(pal) >= 2:
        return True
    if kg and pal:
        return True
    return False


def _is_hybrid_weight_equipment_grid(df: pd.DataFrame) -> bool:
    return _is_ltl_transport_grid(df) and len(_equipment_rate_columns(df)) > 0


def _is_bracket_equipment_hybrid(df: pd.DataFrame) -> bool:
    """Bracket transport (kg/pallet) plus FTL or other equipment columns."""
    if _is_emons_synthetic_min_max_grid(df):
        return True
    if _is_hybrid_weight_equipment_grid(df):
        return True
    if not _bracket_families(df):
        return False
    return len(_equipment_rate_columns(df)) > 0


def _ltl_weight_bands_ordered(df: pd.DataFrame) -> list[str]:
    """Kg-style bands for classic MIN/MAX LTL grids (legacy name)."""
    seen: list[str] = []
    kg_bands = _kg_bracket_columns_ordered(df)
    for val in df["rate_column"]:
        if is_blank(val):
            continue
        text = str(val).strip()
        if text in (LTL_MIN_RATE, LTL_MAX_RATE, LTL_ROUND_TRIP):
            continue
        if _is_min_rate_column(text) or _is_max_rate_column(text):
            continue
        if text in kg_bands:
            continue
        if _ltl_band_header(text) and text not in seen:
            seen.append(text)
    bands = kg_bands + seen
    bands.sort(key=_weight_band_sort_value)
    return bands


def _ltl_price_per_key(price_per: Any) -> str | None:
    if is_blank(price_per):
        return None
    try:
        if pd.isna(price_per):
            return None
    except (TypeError, ValueError):
        pass
    text = str(price_per).strip()
    if normalize_header(text) in ("nan", "none"):
        return None
    return text


def _ltl_price_per_values(df: pd.DataFrame) -> list[str]:
    if "price_per" not in df.columns:
        return []
    by_norm: dict[str, str] = {}
    for val in df["price_per"]:
        key = _ltl_price_per_key(val)
        if key is None or not _is_transport_price_per(key):
            continue
        norm = normalize_header(key)
        if norm not in by_norm:
            by_norm[norm] = key
    return list(by_norm.values())


def _ltl_price_per_values_ordered(df: pd.DataFrame) -> list[str]:
    """Per-unit blocks first, flat/shipment blocks last."""
    values = _ltl_price_per_values(df)
    return sorted(values, key=lambda pp: (1 if is_flat_price_per(pp) else 0, pp))


def _ltl_grouping_cols(
    df: pd.DataFrame,
    shipment_cols: list[tuple[str, str]],
) -> list[str]:
    cols = _shipment_key_cols(shipment_cols)
    if len(_ltl_price_per_values(df)) > 1 and "price_per" in cols:
        cols = [c for c in cols if c != "price_per"]
    return cols


def _charge_label_for_rate_key(df: pd.DataFrame, key: str) -> str:
    if "price_per" not in df.columns:
        return ""
    for _, row in df.iterrows():
        rc = str(row.get("rate_column", "")).strip()
        if _weight_rate_column_key(rc) == key:
            label = row.get("price_per")
            if not is_blank(label):
                return str(label).strip()
    return ""


def _column_unit_suffix(charge_label: str, key: str) -> str:
    if key in (LTL_MIN_RATE, LTL_MAX_RATE):
        return "Flat"
    if is_flat_charge_label(charge_label):
        return "Flat"
    if is_blank(charge_label):
        return "p/unit"
    return format_p_unit_label(charge_label)


def _bracket_cost_specs(
    df: pd.DataFrame,
    bands: list[str],
    family_key: str,
) -> list[tuple[str, str, str]]:
    """(rate key, Excel header, charge label) for one bracket family."""
    specs: list[tuple[str, str, str]] = [("__currency__", "Currency", "")]
    kg_bands = _kg_bracket_columns_ordered(df) if family_key == "kg" else []
    use_min_max = _is_ltl_transport_grid(df) and family_key == "kg"
    if use_min_max:
        specs.append(
            (LTL_MIN_RATE, "MIN", _charge_label_for_rate_key(df, LTL_MIN_RATE))
        )
    for band in bands:
        if family_key == "kg":
            header = _kg_band_display_header(band, kg_bands) or band
        else:
            header = _pallet_band_display_header(band) or band
        specs.append(
            (
                band,
                header,
                _charge_label_for_rate_key(df, _weight_rate_column_key(band)),
            )
        )
    if use_min_max:
        specs.append(
            (LTL_MAX_RATE, "MAX", _charge_label_for_rate_key(df, LTL_MAX_RATE))
        )
    return specs


def _ltl_cost_specs(df: pd.DataFrame) -> list[tuple[str, str, str]]:
    """(rate key, Excel header, charge label for unit suffix) per cost column."""
    bands = _ltl_weight_bands_ordered(df)
    return _bracket_cost_specs(df, bands, "kg")


def _column_has_values(df: pd.DataFrame, internal: str) -> bool:
    if internal not in df.columns:
        return False
    return bool(df[internal].apply(lambda v: not is_blank(v)).any())


def _applies_if_text(rate_column: str) -> str:
    if normalize_header(rate_column) == normalize_header(ROUNDTRIP_RATE_COLUMN):
        return "Applies if: Roundtrip equals True"
    if str(rate_column).endswith(" Roundtrip"):
        equipment = str(rate_column)[: -len(" Roundtrip")]
        return f"Applies if: Equipment type equals {equipment} and Roundtrip equals True"
    return f"Applies if: Equipment type equals {rate_column}"


def _row_is_roundtrip(row: pd.Series) -> bool:
    if "roundtrip" in row.index and is_truthy(row.get("roundtrip")):
        return True
    return is_roundtrip_rate_group(row.get("rate_group"))


def _courier_service_label(rate_group: Any) -> Any:
    """Show Service without One way / Round Trip (those map to cost blocks)."""
    if is_blank(rate_group):
        return rate_group
    parts = [p.strip() for p in str(rate_group).split("|")]
    kept = [
        p
        for p in parts
        if normalize_header(p) not in ("one way", "round trip", "roundtrip")
    ]
    return " | ".join(kept) if kept else parts[0].strip()


def _equipment_rate_columns(df: pd.DataFrame) -> list[str]:
    seen: list[str] = []
    for val in df["rate_column"]:
        if is_blank(val):
            continue
        text = str(val).strip()
        if text in (ROUNDTRIP_RATE_COLUMN, LTL_ROUND_TRIP, "Price"):
            continue
        if _is_maut_rate_column(text):
            continue
        if _is_weight_bracket_rate_column(text):
            continue
        if text not in seen:
            seen.append(text)
    return seen


def _is_ftl_style_rate_group(rate_group: Any) -> bool:
    if is_blank(rate_group):
        return False
    t = normalize_header(str(rate_group))
    return "ftl" in t or "one way" in t or "round trip" in t


def _is_equipment_courier_grid(df: pd.DataFrame) -> bool:
    """Equipment-type price columns with roundtrip carried in Trip Type / Service."""
    if _is_bracket_equipment_hybrid(df):
        return False
    if _is_open_bracket_transport_grid(df):
        return False
    equipment = _equipment_rate_columns(df)
    if not equipment:
        return False
    has_roundtrip_in_service = False
    if "rate_group" in df.columns:
        has_roundtrip_in_service = bool(
            df["rate_group"].astype(str).str.contains("round trip", case=False, na=False).any()
        )
    has_roundtrip_col = False
    if "roundtrip" in df.columns:
        has_roundtrip_col = bool(df["roundtrip"].apply(is_truthy).any())
    return has_roundtrip_in_service or has_roundtrip_col


def _equipment_block_keys(equipment_cols: list[str]) -> list[str]:
    keys: list[str] = []
    for eq in equipment_cols:
        keys.append(eq)
    for eq in equipment_cols:
        keys.append(f"{eq} Roundtrip")
    return keys


def _build_equipment_courier_matrix_rows(
    df: pd.DataFrame,
    shipment_cols: list[tuple[str, str]],
    block_keys: list[str],
    equipment_cols: list[str],
) -> list[dict[str, Any]]:
    key_cols = [
        c for c in _shipment_key_cols(shipment_cols) if c != "rate_group"
    ]
    if not key_cols:
        key_cols = [
            c
            for c in df.columns
            if c not in META_COLUMNS and c not in ("rate_column", "price", "currency")
        ]

    records: list[dict[str, Any]] = []
    grouped = df.groupby(key_cols, dropna=False, sort=False)

    for key, group in grouped:
        if not isinstance(key, tuple):
            key = (key,)
        row_data: dict[str, Any] = {}
        for internal, _ in shipment_cols:
            if internal in key_cols:
                idx = key_cols.index(internal)
                row_data[internal] = key[idx] if idx < len(key) else None
            elif internal == "rate_group":
                row_data[internal] = _courier_service_label(group.iloc[0].get("rate_group"))
            else:
                row_data[internal] = None

        roundtrip_flags = group.apply(_row_is_roundtrip, axis=1)
        prices: dict[str, tuple[Any, Any]] = {}
        for block_key in block_keys:
            is_rt_block = block_key.endswith(" Roundtrip")
            equipment = block_key[: -len(" Roundtrip")] if is_rt_block else block_key
            sub = group[
                (group["rate_column"].astype(str).str.strip() == equipment)
                & (roundtrip_flags == is_rt_block)
            ]
            if sub.empty:
                prices[block_key] = (None, None)
                continue
            first = sub.iloc[0]
            prices[block_key] = (first.get("currency"), first.get("price"))

        if not _has_nonzero_cost(prices):
            continue

        row_data["_prices"] = prices
        records.append(row_data)

    return records


def _export_equipment_courier_matrix(
    df: pd.DataFrame,
    output_path: Path,
    *,
    sheet_name: str,
    include_optional_shipment_columns: bool,
    extra_shipment_columns: OrderedDict[str, str] | None,
) -> Path:
    shipment_cols = _shipment_columns(
        df,
        include_optional=include_optional_shipment_columns,
        extra_columns=extra_shipment_columns,
    )
    equipment_cols = _equipment_rate_columns(df)
    block_keys = _equipment_block_keys(equipment_cols)
    matrix_rows = _build_equipment_courier_matrix_rows(
        df, shipment_cols, block_keys, equipment_cols
    )
    if not matrix_rows:
        raise ValueError("No matrix rows with non-zero equipment courier costs.")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    n_ship = 1 + len(shipment_cols)
    col_cursor = n_ship + 1
    block_starts: dict[str, int] = {}

    lane_header = ws.cell(row=HEADER_ROW_COUNT, column=1, value=LANE_NUM_HEADER)
    lane_header.font = BOLD
    lane_header.fill = HEADER_FILL
    lane_header.alignment = LEFT
    for i, (_, header) in enumerate(shipment_cols, start=2):
        cell = ws.cell(row=HEADER_ROW_COUNT, column=i, value=header)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = LEFT

    for block_key in block_keys:
        block_starts[block_key] = col_cursor
        _write_cost_headers(ws, start_col=col_cursor, rate_column=block_key)
        col_cursor += 2

    data_start_row = HEADER_ROW_COUNT + 1
    for row_offset, rec in enumerate(matrix_rows):
        excel_row = data_start_row + row_offset
        ws.cell(row=excel_row, column=1, value=row_offset + 1)
        for i, (internal, _) in enumerate(shipment_cols, start=2):
            val = rec.get(internal)
            if not is_blank(val):
                ws.cell(row=excel_row, column=i, value=val)

        prices: dict[str, tuple[Any, Any]] = rec.get("_prices", {})
        for block_key in block_keys:
            start = block_starts[block_key]
            currency, price = prices.get(block_key, (None, None))
            if not is_blank(currency):
                ws.cell(row=excel_row, column=start, value=currency)
            if not is_blank(price):
                ws.cell(row=excel_row, column=start + 1, value=price)

    for col_idx in range(1, col_cursor):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 16

    wb.save(output_path)
    return output_path


def _display_rate_group(rate_group: Any) -> Any:
    if _is_ftl_style_rate_group(rate_group):
        return _courier_service_label(rate_group)
    return rate_group


def _extract_bracket_prices(
    group: pd.DataFrame,
    bands: list[str],
    *,
    price_per_norm: str | None = None,
) -> dict[str, Any]:
    prices: dict[str, Any] = {}
    currency: Any = None
    for _, row in group.iterrows():
        rc = str(row["rate_column"]).strip()
        if rc not in bands:
            continue
        if price_per_norm is not None:
            pp = normalize_header(row.get("price_per"))
            if pp != price_per_norm:
                continue
        if currency is None and not is_blank(row.get("currency")):
            currency = row.get("currency")
        prices[_weight_rate_column_key(rc)] = row.get("price")
    prices["__currency__"] = currency
    return prices


def _extract_emons_flat_prices(
    group: pd.DataFrame,
    cost_specs: list[tuple[str, str, str]],
) -> dict[str, Any]:
    """MIN = FLAT RATE 100 kg; MAX = FLAT RATE FTL; flat kg bands in between."""
    prices: dict[str, Any] = {}
    currency: Any = None
    flat_bands = [
        key
        for key, _, _ in cost_specs
        if key not in ("__currency__", LTL_MIN_RATE, LTL_MAX_RATE)
    ]
    for _, row in group.iterrows():
        rc = str(row["rate_column"]).strip()
        pp = normalize_header(row.get("price_per"))
        if pp != "flat rate":
            continue
        if currency is None and not is_blank(row.get("currency")):
            currency = row.get("currency")
        price = row.get("price")
        if rc == EMONS_FLAT_MIN_KG:
            prices[LTL_MIN_RATE] = price
            if rc in flat_bands:
                prices[rc] = price
        elif normalize_header(rc) == normalize_header(EMONS_FLAT_MAX_RC):
            prices[LTL_MAX_RATE] = price
        elif rc in flat_bands:
            prices[rc] = price
    prices["__currency__"] = currency
    return prices


def _extract_minmax_flat_prices(
    group: pd.DataFrame,
    cost_specs: list[tuple[str, str, str]],
) -> dict[str, Any]:
    """MINIMUM / FLAT RATE kg bands / MAXIMUM for classic lane ratebooks (e.g. Gefco)."""
    prices: dict[str, Any] = {}
    currency: Any = None
    flat_bands = [
        key
        for key, _, _ in cost_specs
        if key not in ("__currency__", LTL_MIN_RATE, LTL_MAX_RATE)
    ]
    for _, row in group.iterrows():
        rc = str(row["rate_column"]).strip()
        pp = normalize_header(row.get("price_per"))
        if currency is None and not is_blank(row.get("currency")):
            currency = row.get("currency")
        price = row.get("price")
        if _is_min_rate_column(rc) or pp == "minimum (flat rate)":
            prices[LTL_MIN_RATE] = price
        elif _is_max_rate_column(rc) or pp == "maximum (flat rate)":
            prices[LTL_MAX_RATE] = price
        elif rc in flat_bands and pp == "flat rate":
            prices[rc] = price
    prices["__currency__"] = currency
    return prices


def _extract_full_minmax_ltl_prices(
    group: pd.DataFrame,
    cost_specs: list[tuple[str, str, str]],
) -> dict[str, Any]:
    """Single MIN/MAX block with all kg bands (legacy single-charge export)."""
    prices: dict[str, Any] = {}
    currency: Any = None
    bands = [
        key
        for key, _, _ in cost_specs
        if key not in ("__currency__", LTL_MIN_RATE, LTL_MAX_RATE)
    ]
    for _, row in group.iterrows():
        rc = str(row["rate_column"]).strip()
        pp = normalize_header(row.get("price_per"))
        if currency is None and not is_blank(row.get("currency")):
            currency = row.get("currency")
        price = row.get("price")
        if _is_min_rate_column(rc) or pp == "minimum (flat rate)":
            prices[LTL_MIN_RATE] = price
        elif _is_max_rate_column(rc) or pp == "maximum (flat rate)":
            prices[LTL_MAX_RATE] = price
        elif rc in bands:
            prices[rc] = price
    prices["__currency__"] = currency
    return prices


def _extract_bracket_block_prices(
    group: pd.DataFrame,
    family_key: str,
    block_pp: str,
    cost_specs: list[tuple[str, str, str]],
) -> dict[str, Any]:
    bands = [
        key
        for key, _, _ in cost_specs
        if key not in ("__currency__", LTL_MIN_RATE, LTL_MAX_RATE)
    ]
    if family_key == "kg_combined":
        return _extract_combined_minmax_prices(
            group, cost_specs, block_pp or None
        )
    if family_key == "kg_flat":
        if group["rate_column"].astype(str).apply(_is_min_rate_column).any():
            return _extract_minmax_flat_prices(group, cost_specs)
        return _extract_emons_flat_prices(group, cost_specs)
    if family_key == "pallet":
        return _extract_bracket_prices(group, bands)
    if family_key in ("kg_per100", "kg_groupage"):
        return _extract_bracket_prices(group, bands, price_per_norm="per 100 kg")
    if family_key == "kg":
        return _extract_full_minmax_ltl_prices(group, cost_specs)
    pp_norm = normalize_header(block_pp) if block_pp else None
    return _extract_bracket_prices(group, bands, price_per_norm=pp_norm)


def _build_bracket_equipment_matrix_rows(
    df: pd.DataFrame,
    shipment_cols: list[tuple[str, str]],
    bracket_specs: list[tuple[str, str, list[tuple[str, str, str]]]],
    equipment_block_keys: list[str],
) -> list[dict[str, Any]]:
    """bracket_specs: (family_key, price_per_key, cost_specs) per transport block."""
    key_cols = _shipment_key_cols(shipment_cols)
    if len(_ltl_price_per_values(df)) > 1 and "price_per" in key_cols:
        key_cols = [c for c in key_cols if c != "price_per"]

    records: list[dict[str, Any]] = []
    grouped = df.groupby(key_cols, dropna=False, sort=False)

    for key, group in grouped:
        if not isinstance(key, tuple):
            key = (key,)
        row_data: dict[str, Any] = {}
        for internal, _ in shipment_cols:
            if internal in key_cols:
                idx = key_cols.index(internal)
                val = key[idx] if idx < len(key) else None
                row_data[internal] = (
                    _display_rate_group(val) if internal == "rate_group" else val
                )
            else:
                row_data[internal] = None

        bracket_blocks: dict[tuple[str, str], dict[str, Any]] = {}
        for family_key, block_pp, cost_specs in bracket_specs:
            prices = _extract_bracket_block_prices(
                group, family_key, block_pp, cost_specs
            )
            block_key = (family_key, block_pp)
            if _has_nonzero_ltl_cost(prices, cost_specs):
                bracket_blocks[block_key] = prices

        roundtrip_flags = group.apply(_row_is_roundtrip, axis=1)
        equipment_prices: dict[str, tuple[Any, Any]] = {}
        for block_key in equipment_block_keys:
            is_rt_block = block_key.endswith(" Roundtrip")
            equipment = (
                block_key[: -len(" Roundtrip")] if is_rt_block else block_key
            )
            sub = group[
                (group["rate_column"].astype(str).str.strip() == equipment)
                & (roundtrip_flags == is_rt_block)
            ]
            if sub.empty:
                equipment_prices[block_key] = (None, None)
                continue
            if normalize_header(equipment) == "ftl" and "price_per" in sub.columns:
                flat_sub = sub[
                    sub["price_per"]
                    .astype(str)
                    .apply(lambda x: normalize_header(x) == "flat rate")
                ]
                if not flat_sub.empty:
                    sub = flat_sub
            first = sub.iloc[0]
            equipment_prices[block_key] = (
                first.get("currency"),
                first.get("price"),
            )

        if not bracket_blocks and not _has_nonzero_cost(equipment_prices):
            continue

        row_data["_bracket_blocks"] = bracket_blocks
        row_data["_prices"] = equipment_prices
        records.append(row_data)

    return records


def _export_bracket_equipment_matrix(
    df: pd.DataFrame,
    output_path: Path,
    *,
    sheet_name: str,
    include_optional_shipment_columns: bool,
    extra_shipment_columns: OrderedDict[str, str] | None,
) -> Path:
    shipment_cols = _shipment_columns(
        df,
        include_optional=include_optional_shipment_columns,
        extra_columns=extra_shipment_columns,
    )
    bracket_specs = _bracket_transport_block_specs(df)

    equipment_cols = _equipment_rate_columns(df)
    equipment_block_keys = _equipment_block_keys(equipment_cols)

    matrix_rows = _build_bracket_equipment_matrix_rows(
        df, shipment_cols, bracket_specs, equipment_block_keys
    )
    if not matrix_rows:
        raise ValueError("No matrix rows with non-zero bracket/equipment costs.")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    n_ship = 1 + len(shipment_cols)
    col_cursor = n_ship + 1
    bracket_block_starts: dict[tuple[str, str], int] = {}
    block_cost_specs: dict[tuple[str, str], list[tuple[str, str, str]]] = {}
    equipment_block_starts: dict[str, int] = {}

    lane_header = ws.cell(row=HEADER_ROW_COUNT, column=1, value=LANE_NUM_HEADER)
    lane_header.font = BOLD
    lane_header.fill = HEADER_FILL
    lane_header.alignment = LEFT
    for i, (_, header) in enumerate(shipment_cols, start=2):
        cell = ws.cell(row=HEADER_ROW_COUNT, column=i, value=header)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = LEFT

    for family_key, block_pp, cost_specs in bracket_specs:
        block_key = (family_key, block_pp)
        bracket_block_starts[block_key] = col_cursor
        block_cost_specs[block_key] = cost_specs
        all_flat = family_key in ("pallet", "kg_flat", "kg") or is_flat_charge_label(
            block_pp
        )
        p_unit = block_pp
        if family_key == "kg_combined":
            all_flat = False
            p_unit = "Flat / p/100 kg"
        p_unit_label = "Flat" if all_flat else format_p_unit_label(p_unit)
        if family_key == "kg_combined":
            p_unit_label = "Flat / p/100 kg"
        _write_ltl_transport_headers(
            ws,
            start_col=col_cursor,
            cost_specs=cost_specs,
            block_title=_bracket_block_title(family_key, block_pp),
            p_unit_label=p_unit_label,
            all_flat=all_flat,
            rate_by_category=_bracket_rate_by_category(family_key),
        )
        col_cursor += len(cost_specs)

    for block_key in equipment_block_keys:
        equipment_block_starts[block_key] = col_cursor
        _write_cost_headers(ws, start_col=col_cursor, rate_column=block_key)
        col_cursor += 2

    data_start_row = HEADER_ROW_COUNT + 1
    for row_offset, rec in enumerate(matrix_rows):
        excel_row = data_start_row + row_offset
        ws.cell(row=excel_row, column=1, value=row_offset + 1)
        for i, (internal, _) in enumerate(shipment_cols, start=2):
            val = rec.get(internal)
            if not is_blank(val):
                ws.cell(row=excel_row, column=i, value=val)

        for block_key, prices in rec.get("_bracket_blocks", {}).items():
            start = bracket_block_starts.get(block_key)
            cost_specs = block_cost_specs.get(block_key)
            if start is None or cost_specs is None:
                continue
            for offset, (key, _, _) in enumerate(cost_specs):
                val = prices.get(key)
                if not is_blank(val):
                    ws.cell(row=excel_row, column=start + offset, value=val)

        for block_key in equipment_block_keys:
            start = equipment_block_starts[block_key]
            currency, price = rec.get("_prices", {}).get(block_key, (None, None))
            if not is_blank(currency):
                ws.cell(row=excel_row, column=start, value=currency)
            if not is_blank(price):
                ws.cell(row=excel_row, column=start + 1, value=price)

    for col_idx in range(1, col_cursor):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 16

    wb.save(output_path)
    return output_path


def _export_open_bracket_transport_matrix(
    df: pd.DataFrame,
    output_path: Path,
    *,
    sheet_name: str,
    include_optional_shipment_columns: bool,
    extra_shipment_columns: OrderedDict[str, str] | None,
) -> Path:
    """Kg and/or pallet bracket grids without equipment columns (e.g. LTL Pallets tab only)."""
    shipment_cols = _shipment_columns(
        df,
        include_optional=include_optional_shipment_columns,
        extra_columns=extra_shipment_columns,
    )
    bracket_specs = _bracket_transport_block_specs(df)

    matrix_rows = _build_bracket_equipment_matrix_rows(
        df, shipment_cols, bracket_specs, []
    )
    if not matrix_rows:
        raise ValueError("No matrix rows with non-zero bracket costs.")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    n_ship = 1 + len(shipment_cols)
    col_cursor = n_ship + 1
    bracket_block_starts: dict[tuple[str, str], int] = {}
    block_cost_specs: dict[tuple[str, str], list[tuple[str, str, str]]] = {}

    lane_header = ws.cell(row=HEADER_ROW_COUNT, column=1, value=LANE_NUM_HEADER)
    lane_header.font = BOLD
    lane_header.fill = HEADER_FILL
    lane_header.alignment = LEFT
    for i, (_, header) in enumerate(shipment_cols, start=2):
        cell = ws.cell(row=HEADER_ROW_COUNT, column=i, value=header)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = LEFT

    for family_key, block_pp, cost_specs in bracket_specs:
        block_key = (family_key, block_pp)
        bracket_block_starts[block_key] = col_cursor
        block_cost_specs[block_key] = cost_specs
        all_flat = family_key in ("pallet", "kg_flat", "kg") or is_flat_charge_label(
            block_pp
        )
        p_unit = block_pp
        if family_key == "kg_combined":
            all_flat = False
            p_unit = "Flat / p/100 kg"
        p_unit_label = "Flat" if all_flat else format_p_unit_label(p_unit)
        if family_key == "kg_combined":
            p_unit_label = "Flat / p/100 kg"
        _write_ltl_transport_headers(
            ws,
            start_col=col_cursor,
            cost_specs=cost_specs,
            block_title=_bracket_block_title(family_key, block_pp),
            p_unit_label=p_unit_label,
            all_flat=all_flat,
            rate_by_category=_bracket_rate_by_category(family_key),
        )
        col_cursor += len(cost_specs)

    data_start_row = HEADER_ROW_COUNT + 1
    for row_offset, rec in enumerate(matrix_rows):
        excel_row = data_start_row + row_offset
        ws.cell(row=excel_row, column=1, value=row_offset + 1)
        for i, (internal, _) in enumerate(shipment_cols, start=2):
            val = rec.get(internal)
            if not is_blank(val):
                ws.cell(row=excel_row, column=i, value=val)

        for block_key, prices in rec.get("_bracket_blocks", {}).items():
            start = bracket_block_starts.get(block_key)
            cost_specs = block_cost_specs.get(block_key)
            if start is None or cost_specs is None:
                continue
            for offset, (key, _, _) in enumerate(cost_specs):
                val = prices.get(key)
                if not is_blank(val):
                    ws.cell(row=excel_row, column=start + offset, value=val)

    for col_idx in range(1, col_cursor):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 16

    wb.save(output_path)
    return output_path


def _has_multiple_tabs(df: pd.DataFrame) -> bool:
    if "sheet_name" not in df.columns:
        return False
    names = [
        str(v).strip()
        for v in df["sheet_name"].dropna().unique()
        if not is_blank(v)
    ]
    return len(names) > 1


def _shipment_columns(
    df: pd.DataFrame,
    *,
    include_optional: bool = False,
    extra_columns: OrderedDict[str, str] | None = None,
) -> list[tuple[str, str]]:
    """Return (internal_col, excel_header) pairs to export, in order."""
    maps: list[OrderedDict[str, str]] = [SHIPMENT_COLUMN_MAP]
    if include_optional:
        maps.append(OPTIONAL_SHIPMENT_COLUMN_MAP)
    if extra_columns:
        maps.append(extra_columns)

    seen_headers: set[str] = set()
    result: list[tuple[str, str]] = []
    if _has_multiple_tabs(df):
        result.append(("sheet_name", TAB_HEADER))
        seen_headers.add(TAB_HEADER)
    for mapping in maps:
        for internal, header in mapping.items():
            if internal not in df.columns:
                continue
            if internal in CONDITIONAL_SHIPMENT_COLUMNS and not _column_has_values(df, internal):
                continue
            if header in seen_headers:
                continue
            seen_headers.add(header)
            result.append((internal, header))
    return result


def _shipment_key_cols(shipment_cols: list[tuple[str, str]]) -> list[str]:
    return [internal for internal, _ in shipment_cols]


def _price_is_zero_or_blank(value: Any) -> bool:
    if is_blank(value):
        return True
    try:
        return float(value) == 0
    except (TypeError, ValueError):
        return False


def _has_nonzero_cost(prices: dict[str, tuple[Any, Any]]) -> bool:
    """True if at least one cost block has a non-zero price."""
    return any(not _price_is_zero_or_blank(price) for _currency, price in prices.values())


def _rate_column_order(df: pd.DataFrame) -> list[str]:
    seen: list[str] = []
    for val in df["rate_column"]:
        if is_blank(val):
            continue
        text = str(val).strip()
        if text not in seen:
            seen.append(text)
    if ROUNDTRIP_RATE_COLUMN in seen:
        seen = [c for c in seen if c != ROUNDTRIP_RATE_COLUMN] + [ROUNDTRIP_RATE_COLUMN]
    return seen


def _build_matrix_rows(
    df: pd.DataFrame,
    shipment_cols: list[tuple[str, str]],
    rate_columns: list[str],
) -> list[dict[str, Any]]:
    """One output row per unique shipment; prices keyed by rate_column."""
    key_cols = _shipment_key_cols(shipment_cols)
    if not key_cols:
        key_cols = [
            c
            for c in df.columns
            if c not in META_COLUMNS and c not in ("rate_column", "price", "currency")
        ]

    records: list[dict[str, Any]] = []
    grouped = df.groupby(key_cols, dropna=False, sort=False)

    for key, group in grouped:
        if not isinstance(key, tuple):
            key = (key,)
        row_data: dict[str, Any] = {}
        for i, (internal, _) in enumerate(shipment_cols):
            row_data[internal] = key[i] if i < len(key) else None

        prices: dict[str, tuple[Any, Any]] = {}
        for rate_col in rate_columns:
            sub = group[group["rate_column"].astype(str).str.strip() == rate_col]
            if sub.empty:
                prices[rate_col] = (None, None)
                continue
            first = sub.iloc[0]
            currency = first.get("currency")
            price = first.get("price")
            prices[rate_col] = (currency, price)

        if not _has_nonzero_cost(prices):
            continue

        row_data["_prices"] = prices
        records.append(row_data)

    return records


def _extract_ltl_prices(group: pd.DataFrame) -> tuple[dict[str, Any], Any, Any]:
    """Build price map, currency, and price_per for one shipment group."""
    prices: dict[str, Any] = {}
    currency: Any = None
    price_per: Any = None
    for _, row in group.iterrows():
        rc = str(row["rate_column"]).strip()
        if rc == LTL_ROUND_TRIP:
            continue
        if currency is None and not is_blank(row.get("currency")):
            currency = row.get("currency")
        if price_per is None and not is_blank(row.get("price_per")):
            price_per = row.get("price_per")
        if _is_weight_bracket_rate_column(rc):
            prices[_weight_rate_column_key(rc)] = row.get("price")
    prices["__currency__"] = currency
    return prices, currency, price_per


def _ltl_block_key_for_row(price_per: Any, df: pd.DataFrame) -> str | None:
    values = _ltl_price_per_values(df)
    if not values:
        return ""
    key = _ltl_price_per_key(price_per)
    if key is None:
        return None
    norm = normalize_header(key)
    for v in values:
        if normalize_header(v) == norm:
            return v
    return key


def _build_ltl_matrix_rows(
    df: pd.DataFrame,
    shipment_cols: list[tuple[str, str]],
    cost_specs: list[tuple[str, str]],
) -> list[dict[str, Any]]:
    """One row per shipment; _blocks maps price_per -> price dict."""
    key_cols = _ltl_grouping_cols(df, shipment_cols)
    if not key_cols:
        key_cols = [
            c
            for c in df.columns
            if c not in META_COLUMNS and c not in ("rate_column", "price", "currency")
        ]

    records: list[dict[str, Any]] = []
    grouped = df.groupby(key_cols, dropna=False, sort=False)

    for key, group in grouped:
        if not isinstance(key, tuple):
            key = (key,)
        row_data: dict[str, Any] = {}
        for internal, _ in shipment_cols:
            if internal in key_cols:
                idx = key_cols.index(internal)
                row_data[internal] = key[idx] if idx < len(key) else None
            else:
                row_data[internal] = None

        prices, _currency, price_per = _extract_ltl_prices(group)
        price_per_key = _ltl_block_key_for_row(price_per, df)
        if price_per_key is None:
            continue

        if not _has_nonzero_ltl_cost(prices, cost_specs):
            continue

        row_data["price_per"] = price_per
        row_data["_blocks"] = {price_per_key: prices}
        records.append(row_data)

    return records


def _has_nonzero_ltl_cost(
    prices: dict[str, Any],
    cost_specs: list[tuple[str, str, str]],
) -> bool:
    for key, _header, _charge in cost_specs:
        if key == "__currency__":
            continue
        if not _price_is_zero_or_blank(prices.get(key)):
            return True
    return False


def _write_ltl_transport_headers(
    ws,
    *,
    start_col: int,
    cost_specs: list[tuple[str, str, str]],
    block_title: str,
    p_unit_label: str,
    all_flat: bool,
    rate_by_category: str = "Weight/kg",
) -> None:
    """Single Transport cost block: Currency, MIN, weight bands, MAX."""
    n_cols = len(cost_specs)
    end_col = start_col + n_cols - 1

    ws.merge_cells(
        start_row=1,
        start_column=start_col,
        end_row=1,
        end_column=end_col,
    )
    cell = ws.cell(row=1, column=start_col, value=block_title)
    cell.font = BOLD
    cell.fill = HEADER_FILL
    cell.alignment = LEFT

    ws.merge_cells(
        start_row=2,
        start_column=start_col,
        end_row=2,
        end_column=end_col,
    )
    cell = ws.cell(row=2, column=start_col, value="")
    cell.fill = HEADER_FILL

    ws.merge_cells(
        start_row=3,
        start_column=start_col,
        end_row=3,
        end_column=end_col,
    )
    rate_by = "Flat" if all_flat else (p_unit_label or "p/unit")
    cell = ws.cell(
        row=3,
        column=start_col,
        value=f"Rate by: {rate_by_category} ({rate_by})",
    )
    cell.fill = HEADER_FILL
    cell.alignment = LEFT

    for offset, (key, header, charge_label) in enumerate(cost_specs):
        col = start_col + offset
        if key == "__currency__":
            row4 = header
        else:
            unit = (
                "Flat"
                if all_flat
                else _column_unit_suffix(charge_label, key)
            )
            row4 = f"{header} ({unit})"
        cell = ws.cell(row=HEADER_ROW_COUNT, column=col, value=row4)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = LEFT


def _export_ltl_transport_matrix(
    df: pd.DataFrame,
    output_path: Path,
    *,
    sheet_name: str,
    include_optional_shipment_columns: bool,
    extra_shipment_columns: OrderedDict[str, str] | None,
) -> Path:
    shipment_cols = _shipment_columns(
        df,
        include_optional=include_optional_shipment_columns,
        extra_columns=extra_shipment_columns,
    )
    cost_specs = _ltl_cost_specs(df)
    price_per_values = _ltl_price_per_values_ordered(df)
    if not price_per_values:
        price_per_values = [""]

    matrix_rows = _build_ltl_matrix_rows(df, shipment_cols, cost_specs)
    if not matrix_rows:
        raise ValueError("No matrix rows with non-zero LTL costs.")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    n_ship = 1 + len(shipment_cols)
    col_cursor = n_ship + 1
    block_starts: dict[str, int] = {}

    lane_header = ws.cell(row=HEADER_ROW_COUNT, column=1, value=LANE_NUM_HEADER)
    lane_header.font = BOLD
    lane_header.fill = HEADER_FILL
    lane_header.alignment = LEFT
    for i, (_, header) in enumerate(shipment_cols, start=2):
        cell = ws.cell(row=HEADER_ROW_COUNT, column=i, value=header)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = LEFT

    for pp in price_per_values:
        block_starts[pp] = col_cursor
        all_flat = is_flat_price_per(pp)
        p_unit_label = "Flat" if all_flat else format_p_unit_label(pp)
        _write_ltl_transport_headers(
            ws,
            start_col=col_cursor,
            cost_specs=cost_specs,
            block_title=ltl_transport_block_title(pp),
            p_unit_label=p_unit_label,
            all_flat=all_flat,
        )
        col_cursor += len(cost_specs)

    data_start_row = HEADER_ROW_COUNT + 1
    for row_offset, rec in enumerate(matrix_rows):
        excel_row = data_start_row + row_offset
        ws.cell(row=excel_row, column=1, value=row_offset + 1)
        for i, (internal, _) in enumerate(shipment_cols, start=2):
            val = rec.get(internal)
            if not is_blank(val):
                ws.cell(row=excel_row, column=i, value=val)

        blocks: dict[str, dict[str, Any]] = rec.get("_blocks", {})
        for pp, prices in blocks.items():
            start = block_starts.get(pp)
            if start is None:
                continue
            for offset, (key, _, _) in enumerate(cost_specs):
                val = prices.get(key)
                if not is_blank(val):
                    ws.cell(row=excel_row, column=start + offset, value=val)

    for col_idx in range(1, col_cursor):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 16

    wb.save(output_path)
    return output_path


def _write_cost_headers(
    ws,
    *,
    start_col: int,
    rate_column: str,
) -> None:
    """Write 4-row cost block header for one rate_column (2 columns wide)."""
    col_currency = start_col
    col_value = start_col + 1
    title = _cost_block_title(rate_column)
    is_weight = _is_weight_band(rate_column)
    rate_by = "Weight/Chargeable kg" if is_weight else "per shipment"
    value_header = "p/unit" if is_weight else "Flat"

    row1_text = title
    row2_text = _applies_if_text(rate_column)
    row3_text = f"Rate by: {rate_by}"

    for row_idx, text in enumerate((row1_text, row2_text, row3_text), start=1):
        ws.merge_cells(
            start_row=row_idx,
            start_column=col_currency,
            end_row=row_idx,
            end_column=col_value,
        )
        cell = ws.cell(row=row_idx, column=col_currency, value=text)
        cell.font = BOLD if row_idx == 1 else Font()
        cell.fill = HEADER_FILL
        cell.alignment = LEFT

    ws.cell(row=HEADER_ROW_COUNT, column=col_currency, value="Currency").font = BOLD
    ws.cell(row=HEADER_ROW_COUNT, column=col_currency).fill = HEADER_FILL
    ws.cell(row=HEADER_ROW_COUNT, column=col_value, value=value_header).font = BOLD
    ws.cell(row=HEADER_ROW_COUNT, column=col_value).fill = HEADER_FILL


def export_rates_matrix(
    df: pd.DataFrame,
    output_path: Path | str,
    *,
    sheet_name: str = "rates",
    include_optional_shipment_columns: bool = False,
    extra_shipment_columns: OrderedDict[str, str] | None = None,
) -> Path:
    """
    Write df to an Excel matrix: shipment columns, then cost blocks per rate_column.

    Each cost block uses two columns (Currency + Flat or p/unit) and a 4-row header.
    Prices are placed only in the column matching that row's rate_column.
    """
    if df.empty:
        raise ValueError("Cannot export an empty DataFrame.")

    if "rate_column" not in df.columns or "price" not in df.columns:
        raise ValueError("DataFrame must contain rate_column and price columns.")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    df = _df_for_matrix_export(df)

    if _is_bracket_equipment_hybrid(df):
        return _export_bracket_equipment_matrix(
            df,
            output_path,
            sheet_name=sheet_name,
            include_optional_shipment_columns=include_optional_shipment_columns,
            extra_shipment_columns=extra_shipment_columns,
        )

    if _is_open_bracket_transport_grid(df):
        return _export_open_bracket_transport_matrix(
            df,
            output_path,
            sheet_name=sheet_name,
            include_optional_shipment_columns=include_optional_shipment_columns,
            extra_shipment_columns=extra_shipment_columns,
        )

    if _is_ltl_transport_grid(df):
        return _export_ltl_transport_matrix(
            df,
            output_path,
            sheet_name=sheet_name,
            include_optional_shipment_columns=include_optional_shipment_columns,
            extra_shipment_columns=extra_shipment_columns,
        )

    if _is_equipment_courier_grid(df):
        return _export_equipment_courier_matrix(
            df,
            output_path,
            sheet_name=sheet_name,
            include_optional_shipment_columns=include_optional_shipment_columns,
            extra_shipment_columns=extra_shipment_columns,
        )

    shipment_cols = _shipment_columns(
        df,
        include_optional=include_optional_shipment_columns,
        extra_columns=extra_shipment_columns,
    )
    rate_columns = _rate_column_order(df)
    if not rate_columns:
        raise ValueError("No rate_column values found in DataFrame.")

    matrix_rows = _build_matrix_rows(df, shipment_cols, rate_columns)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    n_ship = 1 + len(shipment_cols)  # col 1 = row number (Lane #)
    cost_start_col = n_ship + 1

    # Shipment headers on row 4
    lane_header = ws.cell(row=HEADER_ROW_COUNT, column=1, value=LANE_NUM_HEADER)
    lane_header.font = BOLD
    lane_header.fill = HEADER_FILL
    lane_header.alignment = LEFT
    for i, (_, header) in enumerate(shipment_cols, start=2):
        cell = ws.cell(row=HEADER_ROW_COUNT, column=i, value=header)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = LEFT

    # Cost block headers
    col_cursor = cost_start_col
    block_starts: dict[str, int] = {}
    for rate_col in rate_columns:
        block_starts[rate_col] = col_cursor
        _write_cost_headers(ws, start_col=col_cursor, rate_column=rate_col)
        col_cursor += 2

    # Data rows
    data_start_row = HEADER_ROW_COUNT + 1
    for row_offset, rec in enumerate(matrix_rows):
        excel_row = data_start_row + row_offset
        ws.cell(row=excel_row, column=1, value=row_offset + 1)
        for i, (internal, _) in enumerate(shipment_cols, start=2):
            val = rec.get(internal)
            if not is_blank(val):
                ws.cell(row=excel_row, column=i, value=val)

        prices: dict[str, tuple[Any, Any]] = rec.get("_prices", {})
        for rate_col in rate_columns:
            start = block_starts[rate_col]
            currency, price = prices.get(rate_col, (None, None))
            if not is_blank(currency):
                ws.cell(row=excel_row, column=start, value=currency)
            if not is_blank(price):
                ws.cell(row=excel_row, column=start + 1, value=price)

    # Column widths
    for col_idx in range(1, col_cursor):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 16

    wb.save(output_path)
    return output_path


def _default_matrix_output_path(converted_path: Path) -> Path:
    """Place matrix file under output/, mirroring processing/<layout>/ if present."""
    stem = converted_path.stem.replace("_converted", "") + "_matrix.xlsx"
    try:
        rel = converted_path.resolve().relative_to(PROCESSING_DIR.resolve())
        if len(rel.parts) > 1:
            return OUTPUT_DIR / rel.parent / stem
    except ValueError:
        pass
    return OUTPUT_DIR / stem


def export_converted_file_to_matrix(
    converted_path: Path | str,
    output_path: Path | str | None = None,
    *,
    sheet_name: str = "rates",
    include_optional_shipment_columns: bool = False,
    extra_shipment_columns: OrderedDict[str, str] | None = None,
) -> Path:
    """Load a *_converted.xlsx (long format) and write matrix layout to output/."""
    converted_path = Path(converted_path)
    df = pd.read_excel(converted_path, sheet_name=sheet_name)
    if output_path is None:
        output_path = _default_matrix_output_path(converted_path)
    return export_rates_matrix(
        df,
        output_path,
        sheet_name=sheet_name,
        include_optional_shipment_columns=include_optional_shipment_columns,
        extra_shipment_columns=extra_shipment_columns,
    )


def _prompt(message: str) -> str:
    try:
        return input(message).strip()
    except (EOFError, KeyboardInterrupt):
        print("\nCancelled.")
        sys.exit(0)


def _find_converted_files() -> list[Path]:
    if not PROCESSING_DIR.is_dir():
        return []
    return sorted(
        p
        for p in PROCESSING_DIR.rglob("*_converted.xlsx")
        if p.is_file() and not p.name.startswith("~$")
    )


def _choose_converted_file(files: list[Path]) -> Path | None:
    print(f"\nConverted files in {PROCESSING_DIR.name}/:")
    for i, path in enumerate(files, 1):
        rel = path.relative_to(PROJECT_ROOT)
        print(f"  {i}. {rel}")
    print("  0. Exit")

    while True:
        choice = _prompt("\nSelect file number: ")
        if choice == "0":
            return None
        if choice.isdigit() and 1 <= int(choice) <= len(files):
            return files[int(choice) - 1]
        print("Invalid choice. Enter a number from the list.")


def main() -> None:
    print("=" * 60)
    print("  Aptiv Road — matrix Excel export")
    print("=" * 60)
    print(
        "\nInput: a *_converted.xlsx from processing/ (run convert.py first)."
    )
    print(f"Output: {OUTPUT_DIR.name}/<layout>/*_matrix.xlsx\n")

    files = _find_converted_files()
    if not files:
        print(f"No *_converted.xlsx files found under {PROCESSING_DIR}/")
        print("Run:  python convert.py")
        sys.exit(1)

    selected = _choose_converted_file(files)
    if selected is None:
        print("Goodbye.")
        return

    print(f"\nExporting: {selected.name}")
    try:
        out = export_converted_file_to_matrix(selected)
    except Exception as exc:
        print(f"ERROR: {exc}")
        sys.exit(1)

    print(f"Saved -> {out.relative_to(PROJECT_ROOT)}")


if __name__ == "__main__":
    main()
