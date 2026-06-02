"""Shared utilities for rate workbook converters."""

from __future__ import annotations

import re
from collections.abc import Callable
from pathlib import Path
from typing import Any, Iterator

import pandas as pd

SKIP_TAB_RE = re.compile(
    r"revision|general\s*rules|instruction|accessorial|accesorial|"
    r"fsc|fuel\s*surcharge|source|validat|flags|definition\s+and\s+rules|"
    r"^sheet\d+$|tt\s+details|other\s+information|other\s+rates",
    re.I,
)

LAYOUT3_SKIP_EXTRA = re.compile(r"definition|flags|validations?", re.I)


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip().lower()
    return re.sub(r"\s+", " ", text)


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def parse_city_zip_country(value: Any) -> tuple[Any, Any, Any]:
    """
    Split a combined location like 'Grosspetersdorf, 75, AT' into
    (city, postal_code, country). Returns (value, None, None) when not parseable.
    """
    if is_blank(value):
        return None, None, None
    text = str(value).strip()
    parts = [p.strip() for p in text.split(",")]
    if len(parts) < 3:
        return text, None, None
    country = parts[-1]
    postal = parts[-2]
    city = ", ".join(parts[:-2]).strip()
    if not city:
        return text, None, None
    return city, postal, country


def format_p_unit_label(price_per: Any) -> str:
    """
    Build a per-unit label from Price Per, e.g. '100Kg' -> 'p/100 kg'.
    """
    if is_blank(price_per):
        return "p/unit"
    text = str(price_per).strip()
    m = re.match(r"^(\d+(?:[.,]\d+)?)\s*([a-zA-Z]+)$", text.replace(" ", ""))
    if m:
        num = m.group(1).replace(",", "")
        unit = m.group(2).lower()
        return f"p/{num} {unit}"
    m_num = re.match(r"^(\d+(?:[.,]\d+)?)$", text.replace(" ", ""))
    if m_num:
        return f"p/{m_num.group(1).replace(',', '')} unit"
    t = normalize_header(text)
    if t in ("flat rate", "flat"):
        return "Flat"
    if t in ("shipment", "trip", "lane"):
        return "p/shipment"
    per_kg = re.search(r"per\s+(\d+(?:[.,\d]+)?)\s*kg", t)
    if per_kg:
        return f"p/{per_kg.group(1).replace(',', '')} kg"
    return f"p/{text}"


def is_flat_charge_label(value: Any) -> bool:
    if is_blank(value):
        return False
    t = normalize_header(str(value).strip())
    return t in (
        "flat rate",
        "flat",
        "minimum (flat rate)",
        "maximum (flat rate)",
        "min price",
        "max price",
    )


def is_roundtrip_trip(value: Any) -> bool:
    if is_blank(value):
        return False
    t = normalize_header(str(value).strip())
    return "round" in t and ("trip" in t or t == "roundtrip")


def is_roundtrip_rate_group(value: Any) -> bool:
    if is_blank(value):
        return False
    t = normalize_header(str(value).strip())
    return "round trip" in t or t.endswith("roundtrip") or t == "roundtrip"


def is_flat_price_per(price_per: Any) -> bool:
    if is_blank(price_per):
        return False
    return normalize_header(str(price_per).strip()) in ("shipment", "trip", "lane")


def ltl_transport_block_title(price_per: Any) -> str:
    """Excel row-1 title for an LTL / groupage transport block."""
    if is_blank(price_per):
        return "Transport cost"
    if is_flat_price_per(price_per):
        return "Transport cost (Flat)"
    return f"Transport cost ({format_p_unit_label(price_per)})"


def is_truthy(value: Any) -> bool:
    if is_blank(value):
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    return normalize_header(value) in ("true", "yes", "1", "y", "x")


def enrich_location_fields(
    city: Any,
    postal: Any,
    country: Any,
) -> tuple[Any, Any, Any]:
    """Fill missing postal/country (and city) from a combined city field."""
    parsed_city, parsed_postal, parsed_country = parse_city_zip_country(city)
    out_city = parsed_city if parsed_postal is not None else city
    out_postal = postal if not is_blank(postal) else parsed_postal
    out_country = country if not is_blank(country) else parsed_country
    return out_city, out_postal, out_country


NON_PRICE_COLUMN_RE = re.compile(
    r"transit\s*time|time\s+d2d|time\s+d\s*>?\s*d|in\s+hours|in\s+hrs|"
    r"^hrs?$|comment|departure\s+day|arrival\s+date|maximum\s+transit|"
    r"rate\s+currency|^\s*origin\s*$|^\s*destination\s*$|"
    r"lane\s*#|lane\s*id|lane\s*number|paid\s+by|plant\s+code|"
    r"per\s+100\s*kg|minimum\s+charge|^cost\s*$|allocation",
    re.I,
)


def format_cell_as_displayed(value: Any, number_format: str | None = None) -> Any:
    """Format a cell value the way it typically appears in Excel."""
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, bool):
        return value
    if not isinstance(value, (int, float)):
        return value

    fmt = (number_format or "General").split(";")[0].strip()
    if fmt in ("General", "0", "@", ""):
        if float(value).is_integer():
            return str(int(round(float(value))))
        text = f"{float(value):g}"
        return text

    decimals = 0
    use_thousands = "#,##" in fmt or "#.##" in fmt or ",#" in fmt
    if "." in fmt:
        after_dot = fmt.split(".", 1)[1]
        decimals = sum(1 for ch in after_dot if ch in "0#?")

    rounded = round(float(value), decimals)
    if decimals == 0:
        rounded = int(round(float(value)))
        return f"{rounded:,}" if use_thousands else str(rounded)

    text = f"{rounded:,.{decimals}f}" if use_thousands else f"{rounded:.{decimals}f}"
    return text


def forward_fill_row_labels(row: list[Any]) -> list[str]:
    """Forward-fill merged header cells across columns."""
    labels: list[str] = []
    current = ""
    for cell in row:
        if not is_blank(cell):
            current = str(cell).replace("\n", " ").strip()
        labels.append(current)
    return labels


def read_sheet_rows(path: Path, sheet_name: str, *, as_displayed: bool = False) -> list[list[Any]]:
    """Read a worksheet into a list of rows (values only, or display-formatted)."""
    ext = path.suffix.lower()
    if ext == ".xlsx":
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            for name in wb.sheetnames:
                if name.strip().lower() == sheet_name.strip().lower():
                    sheet_name = name
                    break
        ws = wb[sheet_name]
        if not as_displayed:
            rows = [list(row) for row in ws.iter_rows(values_only=True)]
        else:
            rows = []
            for row in ws.iter_rows():
                out_row: list[Any] = []
                for cell in row:
                    out_row.append(
                        format_cell_as_displayed(cell.value, cell.number_format)
                    )
                rows.append(out_row)
        wb.close()
        if rows:
            max_cols = max(len(r) for r in rows)
            rows = [list(r) + [None] * (max_cols - len(r)) for r in rows]
        return rows
    if ext == ".xls":
        import xlrd

        wb = xlrd.open_workbook(path, formatting_info=as_displayed)
        for name in wb.sheet_names():
            if name.strip().lower() == sheet_name.strip().lower():
                sheet_name = name
                break
        sh = wb.sheet_by_name(sheet_name)
        rows = []
        for r in range(sh.nrows):
            row_vals = []
            for c in range(sh.ncols):
                val = sh.cell_value(r, c)
                if as_displayed and sh.cell_type(r, c) == xlrd.XL_CELL_NUMBER:
                    xf = sh.cell_xf_index(r, c)
                    fmt_key = wb.xf_list[xf].format_key
                    fmt = wb.format_map.get(fmt_key)
                    fmt_str = fmt.format_str if fmt else None
                    row_vals.append(format_cell_as_displayed(val, fmt_str))
                else:
                    row_vals.append(val)
            rows.append(row_vals)
        if rows:
            max_cols = max(len(r) for r in rows)
            rows = [list(r) + [None] * (max_cols - len(r)) for r in rows]
        return rows
    raise ValueError(f"Unsupported file type: {path}")


def get_sheet_names(path: Path) -> list[str]:
    return list(iter_sheets(path))


def resolve_sheet_names(path: Path, requested: list[str]) -> list[str]:
    """Map user-provided tab names to actual workbook sheet names."""
    all_sheets = get_sheet_names(path)
    by_lower = {name.strip().lower(): name for name in all_sheets}
    resolved: list[str] = []
    missing: list[str] = []
    for req in requested:
        key = req.strip().lower()
        if key in by_lower:
            resolved.append(by_lower[key])
        else:
            missing.append(req)
    if missing:
        raise ValueError(
            f"Unknown sheet(s): {', '.join(missing)}. "
            f"Available: {', '.join(all_sheets)}"
        )
    return resolved


def sheets_to_convert(
    path: Path,
    *,
    sheets: list[str] | None,
    auto_include: Callable[[str], bool],
) -> list[str]:
    """Return sheet names to process: explicit list or auto-detected price tabs."""
    if sheets is None:
        return [name for name in get_sheet_names(path) if auto_include(name)]
    return resolve_sheet_names(path, sheets)


def iter_sheets(path: Path) -> Iterator[str]:
    ext = path.suffix.lower()
    if ext == ".xlsx":
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        names = list(wb.sheetnames)
        wb.close()
    elif ext == ".xls":
        import xlrd

        wb = xlrd.open_workbook(path, on_demand=True)
        names = wb.sheet_names()
        wb.release_resources()
    else:
        raise ValueError(f"Unsupported file type: {path}")
    yield from names


def should_skip_tab(sheet_name: str, layout: str) -> bool:
    if SKIP_TAB_RE.search(sheet_name):
        return True
    if layout in ("new_grid", "layout3") and LAYOUT3_SKIP_EXTRA.search(sheet_name):
        if "rate grid" in sheet_name.lower():
            return False
        return True
    return False


def output_path(processing_dir: Path, layout: str, source_file: Path) -> Path:
    out_dir = processing_dir / layout
    out_dir.mkdir(parents=True, exist_ok=True)
    return out_dir / f"{source_file.stem}_converted.xlsx"


def reorder_converted_df(
    df: pd.DataFrame,
    data_columns: tuple[str, ...],
) -> pd.DataFrame:
    """Ensure standard data columns exist and appear in a stable order after metadata."""
    meta = ("layout", "source_file", "sheet_name")
    out = df.copy()
    for col in data_columns:
        if col not in out.columns:
            out[col] = None
    ordered = [c for c in meta if c in out.columns]
    ordered += [c for c in data_columns if c in out.columns]
    ordered += [c for c in out.columns if c not in ordered]
    return out[ordered]


def save_dataframe(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(path, index=False, sheet_name="rates")


def append_excel_sheet(
    path: Path,
    df: pd.DataFrame,
    sheet_name: str,
    *,
    apply_accessorial_format: bool = False,
) -> None:
    """Add or replace a sheet in an existing Excel workbook."""
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.is_file():
        with pd.ExcelWriter(
            path,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace",
        ) as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        df.to_excel(path, index=False, sheet_name=sheet_name)
    if apply_accessorial_format:
        format_accessorial_sheet(path, sheet_name)


def format_accessorial_sheet(path: Path, sheet_name: str) -> None:
    """Apply header styling and column widths to an accessorial sheet."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return
    ws = wb[sheet_name]
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    widths = {
        "Rate Card Name": 52,
        "Rate Agreement Name": 18,
        "cost": 14,
        "currency": 10,
        "condition": 28,
        "measurement": 22,
    }
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        title = str(cell.value) if cell.value else ""
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = widths.get(title, 18)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    wb.save(path)
    wb.close()


def copy_excel_sheet(
    source_path: Path,
    dest_path: Path,
    sheet_name: str,
) -> bool:
    """Copy a sheet from source workbook into dest workbook (replace if exists)."""
    if not source_path.is_file():
        return False
    try:
        src_df = pd.read_excel(source_path, sheet_name=sheet_name)
    except ValueError:
        return False
    append_excel_sheet(dest_path, src_df, sheet_name)
    return True


def add_metadata(
    df: pd.DataFrame,
    *,
    source_file: str,
    layout: str,
    sheet_name: str,
) -> pd.DataFrame:
    out = df.copy()
    out.insert(0, "layout", layout)
    out.insert(1, "source_file", source_file)
    out.insert(2, "sheet_name", sheet_name)
    return out


def rows_to_dataframe(rows: list[list[Any]]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame()
    max_cols = max(len(r) for r in rows)
    padded = [list(r) + [None] * (max_cols - len(r)) for r in rows]
    return pd.DataFrame(padded)


def find_row_with_keywords(
    rows: list[list[Any]], keywords: list[str], max_scan: int = 35
) -> int | None:
    keys = [k.lower() for k in keywords]
    for idx in range(min(max_scan, len(rows))):
        text = " ".join(normalize_header(c) for c in rows[idx])
        if all(k in text for k in keys):
            return idx
    return None


def col_index_by_keywords(header_row: list[Any], *keyword_groups: str) -> int | None:
    for idx, cell in enumerate(header_row):
        text = normalize_header(cell)
        for kw in keyword_groups:
            if kw in text:
                return idx
    return None
